"""
Script standalone: extrae concesionarios de autofin.cl y los guarda en .xlsx

Los datos vienen del JSON embebido en el HTML (__NEXT_DATA__ de Next.js),
sin necesidad de parsear el DOM ni usar selectores CSS frágiles.

Uso:
    python scripts/concesionarios_autofin.py

Requiere Playwright y openpyxl:
    .venv\Scripts\pip install openpyxl
    .venv\Scripts\playwright install chromium
"""

import asyncio
import json
import re
from dataclasses import dataclass, fields, astuple

from playwright.async_api import async_playwright

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    raise SystemExit("Falta openpyxl. Instala con: pip install openpyxl")


URL = "https://autofin.cl/concesionarios"
OUTPUT = "concesionarios_autofin.xlsx"


@dataclass
class Concesionario:
    nombre: str
    region: str
    comuna: str
    direccion: str
    horario_semana: str
    horario_finde: str
    web: str
    latitud: str
    longitud: str


async def extraer_concesionarios() -> list[Concesionario]:
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            locale="es-CL",
        )
        page = await context.new_page()

        print(f"Cargando {URL} ...")
        await page.goto(URL, wait_until="networkidle", timeout=30_000)

        html = await page.content()
        await browser.close()

    # Extraer JSON embebido por Next.js
    match = re.search(r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>', html, re.DOTALL)
    if not match:
        raise RuntimeError("No se encontró __NEXT_DATA__ en el HTML. El sitio puede haber cambiado.")

    data = json.loads(match.group(1))
    raw = data["props"]["pageProps"]["concessionaires"]
    print(f"Concesionarios encontrados: {len(raw)}")

    urls_vistas: set[str] = set()
    resultados: list[Concesionario] = []

    for item in raw:
        web = item.get("web", "").strip()

        # Descartar sin web real
        if not web or web in ("javascript:void(0);", "javascript:void(0)", "#"):
            continue

        # Descartar duplicados por URL
        if web in urls_vistas:
            continue
        urls_vistas.add(web)

        resultados.append(Concesionario(
            nombre=item.get("name", "").strip(),
            region=item.get("region", {}).get("name", "").strip(),
            comuna=item.get("commune", {}).get("name", "").strip(),
            direccion=item.get("address", "").strip(),
            horario_semana=item.get("hours", "").strip(),
            horario_finde=item.get("weekendSchedule", "").strip(),
            web=web,
            latitud=item.get("latitude", "").strip(),
            longitud=item.get("longitude", "").strip(),
        ))

    return resultados


def guardar_xlsx(concesionarios: list[Concesionario], ruta: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Concesionarios"

    headers = [f.name.replace("_", " ").capitalize() for f in fields(Concesionario)]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, c in enumerate(concesionarios, 2):
        for col, valor in enumerate(astuple(c), 1):
            ws.cell(row=row, column=col, value=valor)

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    wb.save(ruta)
    print(f"Guardado: {ruta} ({len(concesionarios)} concesionarios)")


async def main() -> None:
    concesionarios = await extraer_concesionarios()
    if not concesionarios:
        print("No se encontraron concesionarios con web válida.")
        return
    guardar_xlsx(concesionarios, OUTPUT)


if __name__ == "__main__":
    asyncio.run(main())
